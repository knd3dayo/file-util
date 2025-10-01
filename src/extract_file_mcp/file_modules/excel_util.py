from typing import Union
import datetime
import json
import openpyxl # type: ignore

class ExcelUtil:

    excel_request_name = "excel_request"
    @classmethod
    def get_excel_request_objects(cls, request_dict: dict) -> tuple[str, dict]:
        '''
        {"context": {"excel_request": {}}}の形式で渡される
        '''
        # contextを取得
        request:Union[dict, None] = request_dict.get(cls.excel_request_name, None)
        if not request:
            raise ValueError("request is not set.")
        # file_pathとdata_jsonを取得
        file_path = request.get("file_path", None)
        data_json = request.get("data_json", "[]")
        data = json.loads(data_json)

        return file_path, data

    @classmethod
    def get_sheet_names_api(cls, request_json: str):
        # request_jsonからrequestを作成
        request_dict: dict = json.loads(request_json)
        # excel_requestを取得
        file_path, _ = ExcelUtil.get_excel_request_objects(request_dict)
        text = ExcelUtil.get_sheet_names(file_path)
        return {"output": text}

    @classmethod
    def extract_excel_sheet_api(cls, request_json: str):
        # request_jsonからrequestを作成
        request_dict: dict = json.loads(request_json)
        # excel_requestを取得
        file_path, excel_request = ExcelUtil.get_excel_request_objects(request_dict)
        # excel_sheet_nameを取得
        sheet_name = excel_request.get("excel_sheet_name", "")

        text = ExcelUtil.extract_text_from_sheet(file_path, sheet_name)
        return {"output": text}

    @classmethod
    def export_to_excel_api(cls, request_json: str):
        # request_jsonからrequestを作成
        request_dict: dict = json.loads(request_json)
        
        # file_pathとdata_jsonを取得
        file_path, dataJson = ExcelUtil.get_excel_request_objects(request_dict)
        ExcelUtil.export_to_excel(file_path, dataJson.get("rows",[]))
        # 結果用のdictを生成
        return {}

    @classmethod
    def import_from_excel_api(cls, request_json: str):
        # request_jsonからrequestを作成
        request_dict: dict = json.loads(request_json)
        # file_requestを取得
        file_path, _ = cls.get_excel_request_objects(request_dict)
        # import_to_excelを実行
        data = ExcelUtil.import_from_excel(file_path)
        # 結果用のdictを生成
        result = {}
        result["rows"] = data
        return result


    @classmethod
    def export_to_excel(cls, filePath, data):
        # Workbookオブジェクトを生成
        wb = openpyxl.Workbook()
        # アクティブなシートを取得
        ws = wb.active
        # シート名を設定
        if ws is None:
            ws = wb.create_sheet()

        ws.title = "Sheet1"
        # データを書き込む
        for row in data:
            ws.append(row)
        # ファイルを保存
        wb.save(filePath)
        
    @classmethod
    def import_from_excel(cls, filePath):
        # Workbookオブジェクトを生成
        wb = openpyxl.load_workbook(filePath)
        # アクティブなシートを取得
        ws = wb.active
        # データを取得
        data = []
        if ws is not None:
            for row in ws.iter_rows(values_only=True):
                data.append(row)
        
        return data

    # application/vnd.openxmlformats-officedocument.spreadsheetml.sheetのファイルを読み込んで文字列として返す関数
    @classmethod
    def extract_text_from_sheet(cls, filename:str, sheet_name:str=""):
        import openpyxl
        from io import StringIO
        # 出力用のストリームを作成
        output = StringIO()
        wb = openpyxl.load_workbook(filename)
        for sheet in wb:
            # シート名が指定されている場合はそのシートのみ処理
            if sheet_name and sheet.title != sheet_name:
                continue
            for row in sheet.iter_rows(values_only=True):
                # 1行分のデータを格納するリスト
                cells = []
                for cell in row:
                    # cell.valueがNoneの場合はcontinue
                    if cell is None:
                        continue
                    # cell.valueがdatetime.datetimeの場合はisoformat()で文字列に変換
                    if isinstance(cell, datetime.datetime):
                        cells.append(cell.isoformat())
                    else:
                        cells.append(str(cell))
                    
                output.write("\t".join(cells))
                output.write("\n")
        
        return output.getvalue()

    # excelのシート名一覧を取得する関数
    @classmethod
    def get_sheet_names(cls, filename):
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        return wb.sheetnames

