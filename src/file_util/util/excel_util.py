import datetime
import openpyxl
from io import StringIO

class ExcelUtil:

    # application/vnd.openxmlformats-officedocument.spreadsheetml.sheetのファイルを読み込んで文字列として返す関数
    @classmethod
    def extract_text_from_sheet(cls, filename:str, sheet_name:str=""):
        # 出力用のストリームを作成
        output = StringIO()
        wb = openpyxl.load_workbook(filename)
        for sheet in wb:
            # シート名が指定されている場合はそのシートのみ処理
            if sheet_name and sheet.title != sheet_name:
                continue
            for row in sheet.iter_rows(values_only=True):
                # 1行分のデータを格納するリスト
                cells = []
                for cell in row:
                    # cell.valueがNoneの場合はcontinue
                    if cell is None:
                        continue
                    # cell.valueがdatetime.datetimeの場合はisoformat()で文字列に変換
                    if isinstance(cell, datetime.datetime):
                        cells.append(cell.isoformat())
                    else:
                        cells.append(str(cell))
                    
                output.write("\t".join(cells))
                output.write("\n")
        
        return output.getvalue()

    # excelのシート名一覧を取得する関数
    @classmethod
    def get_sheet_names(cls, filename):
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        return wb.sheetnames

    # データをExcelファイルにエクスポートする関数
    @classmethod
    def export_data_to_excel(cls, data: dict[str, list], filename, sheet_name: str| None ="Sheet1"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        if sheet_name is not None:
            ws.title = sheet_name

        # ヘッダー行の追加
        headers = list(data.keys())
        ws.append(headers)
        # データ行の追加
        num_rows = len(next(iter(data.values()), []))
        for i in range(num_rows):
            row = []
            for header in headers:
                column_data = data.get(header, [])
                if i < len(column_data):
                    row.append(column_data[i])
                else:
                    row.append("")
            ws.append(row)
        
        wb.save(filename)

    # Excelファイルの内容を辞書型で取得する関数
    @classmethod
    def import_data_from_excel(cls, filename, sheet_name: str | None ="Sheet1") -> dict[str, list]:
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return {}
        
        data: dict[str, list] = {}
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return data

        headers = rows[0]
        for header in headers:
            data[str(header)] = []
        for row in rows[1:]:
            for header, cell in zip(headers, row):
                data[str(header)].append(cell)
        
        return data          